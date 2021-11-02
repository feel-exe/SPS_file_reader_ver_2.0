# pip install -r req.txt

import openpyxl
import pandas as pd
from config.config import columns_min, data_description
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)


class ConvertTXTToExcel():
    def __init__(self, file_path, file_name, columns_min, file_extension):
        self.file_path = file_path
        self.file_name = file_name
        self.columns_min = columns_min
        self.file_extension = file_extension

    def load_txt(self):
        # достаем параметры записи
        head_title = []
        with open(f"{self.file_path}{self.file_name}.{self.file_extension[0].upper()}", encoding='cp1251') as file:
            lines = file.readlines()
        head_title.append(lines[0].split())
        head_title.append(lines[1].split())
        self.set_parameters = lines[1].split()

        # создаем dataframe чистых данных
        self.data = pd.read_csv(f"{self.file_path}{self.file_name}.{self.file_extension[0].upper()}", encoding='cp1251', sep=' ', header=None,
                                skiprows=[0, 1])

        self.data = self.data.drop([7], axis=1)

        self.data.columns = self.columns_min

        for col in columns_min:
            arr = []
            for item in self.data[f"{col}"]:
                arr.append(float(item.split()[-1].replace(",", ".")))
            self.data[f"{col}"] = arr

    def create_exls(self):
        #  создание объекта эксель
        book = openpyxl.Workbook()
        book.remove(book.active)
        sheet = book.create_sheet("Режим спекания")
        sheet = book.active

        #  заполнение датафреймом

        for r in dataframe_to_rows(self.data, index=False, header=True):
            sheet.append(r)

        sheet.insert_rows(0)
        sheet.insert_cols(0)
        sheet.insert_cols(0)
        sheet.insert_cols(0)

        sheet["B2"].value = "t"
        #  заполняем временную шкалу
        i = 1
        # for item in range(0, int(data.shape[0]), int(float(set_parameters[0].replace(",", ".")))):
        for item in range(0, int(self.data.shape[0])):
            sheet[f"B{i + 2}"].value = item * int(float(self.set_parameters[0].replace(",", ".")))
            i += 1
            # sheet.iter_cols(min_col=2, min_row = 2)

        # табличка с описанием

        for item in range(len(data_description)):
            sheet.merge_cells(f'M{29 + item}:P{29 + item}')
            sheet[f'M{29 + item}'].value = data_description[item]

        # MAX
        sheet['M40'].value = "Тмакс"
        # sheet['N40'].value = max(sheet[f'F2:F{2+data.shape[0]}'].value)

        # Temperature
        c1 = ScatterChart()

        xvalues = Reference(sheet, min_col=2, min_row=2, max_row=2 + int(self.data.shape[0]))
        values = Reference(sheet, min_col=6, min_row=2, max_row=2 + int(self.data.shape[0]))
        series = Series(values, xvalues, title_from_data=True)
        c1.series.append(series)

        s1 = c1.series[0]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "0000FF"  # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "0000FF"
        # Marker outline
        s1.graphicalProperties.line.noFill = True  # hide lines

        c1.style = 13
        c1.x_axis.title = 't,сек'
        c1.y_axis.title = 'T, C'
        c1.y_axis.majorGridlines = None
        c1.x_axis.majorGridlines = None

        # Displaciment Rate
        c2 = ScatterChart()

        xvalues = Reference(sheet, min_col=2, min_row=2, max_row=2 + int(self.data.shape[0]))
        values = Reference(sheet, min_col=9, min_row=2, max_row=2 + int(self.data.shape[0]))
        series = Series(values, xvalues, title_from_data=True)
        c2.series.append(series)

        s2 = c2.series[0]
        s2.marker.symbol = "diamond"
        s2.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
        s2.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
        s2.graphicalProperties.line.noFill = True  # hide lines

        c2.style = 13
        # c2.x_axis.title = 'Days'
        c2.y_axis.title = 'S, мм/с'
        c2.y_axis.axId = 200
        c2.y_axis.majorGridlines = None
        c2.x_axis.majorGridlines = None

        # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
        c1.y_axis.crosses = "max"
        c1 += c2

        sheet.add_chart(c1, "O10")

        book.save(f"{self.file_path}{self.file_name}_converted.xlsx")

# file_path = '/home/user/PycharmProjects/SPS_file_reader_ver_2.0/vlt_txt_excel/'
# file_name = "27"

# gg = ConvertTXTToExcel(file_path, file_name, columns_min)
# gg.load_txt()
# gg.create_exls()

#     def read_input_file():
#     # достаем параметры записи
#     head_title = []
#     with open("vlt_txt_excel/27.TXT", encoding='cp1251') as file:
#         lines = file.readlines()
#     head_title.append(lines[0].split())
#     head_title.append(lines[1].split())
#     parameters = lines[1].split()
#
#     # создаем dataframe чистых данных
#     # data = pd.read_csv("vlt_txt_excel/27.TXT", encoding='cp1251', sep='\t', header=None, skiprows=[0, 1],
#     #                      engine='python')
#     data = pd.read_csv("vlt_txt_excel/27.TXT", encoding='cp1251', sep=' ', header=None, skiprows=[0, 1])
#
#     data = data.drop([7], axis=1)
#
#     data.columns = columns_min
#
#     for col in columns_min:
#         arr = []
#         for item in data[f"{col}"]:
#             arr.append(float(item.split()[-1].replace(",", ".")))
#         data[f"{col}"] = arr
#
#     return data, parameters, head_title
#
#
# data, set_parameters, head_title = read_input_file()
#
#
# def create_exls(data):
#     #  создание объекта эксель
#     book = openpyxl.Workbook()
#     book.remove(book.active)
#     sheet = book.create_sheet("Режим спекания")
#     sheet = book.active
#
#
#     #  заполнение датафреймом
#
#     for r in dataframe_to_rows(data, index=False, header=True):
#         sheet.append(r)
#
#     sheet.insert_rows(0)
#     sheet.insert_cols(0)
#     sheet.insert_cols(0)
#     sheet.insert_cols(0)
#
#     sheet["B2"].value = "t"
#     #  заполняем временную шкалу
#     i = 1
#     # for item in range(0, int(data.shape[0]), int(float(set_parameters[0].replace(",", ".")))):
#     for item in range(0, int(data.shape[0])):
#         sheet[f"B{i + 2}"].value = item * int(float(set_parameters[0].replace(",", ".")))
#         i += 1
#         # sheet.iter_cols(min_col=2, min_row = 2)
#
#     # табличка с описанием
#
#     for item in range(len(data_description)):
#         sheet.merge_cells(f'M{29+item}:P{29+item}')
#         sheet[f'M{29+item}'].value = data_description[item]
#
#
#     # MAX
#     sheet['M40'].value = "Тмакс"
#     # sheet['N40'].value = max(sheet[f'F2:F{2+data.shape[0]}'].value)
#
#
#
#     # Temperature
#     c1 = ScatterChart()
#
#     xvalues = Reference(sheet, min_col=2, min_row=2, max_row=2 + int(data.shape[0]))
#     values = Reference(sheet, min_col=6, min_row=2, max_row=2 + int(data.shape[0]))
#     series = Series(values, xvalues, title_from_data=True)
#     c1.series.append(series)
#
#     s1 = c1.series[0]
#     s1.marker.symbol = "triangle"
#     s1.marker.graphicalProperties.solidFill = "0000FF"  # Marker filling
#     s1.marker.graphicalProperties.line.solidFill = "0000FF"
#     # Marker outline
#     s1.graphicalProperties.line.noFill = True  # hide lines
#
#     c1.style = 13
#     c1.x_axis.title = 't,сек'
#     c1.y_axis.title = 'T, C'
#     c1.y_axis.majorGridlines = None
#     c1.x_axis.majorGridlines = None
#
#     # Displaciment Rate
#     c2 = ScatterChart()
#
#     xvalues = Reference(sheet, min_col=2, min_row=2, max_row=2 + int(data.shape[0]))
#     values = Reference(sheet, min_col=9, min_row=2, max_row=2 + int(data.shape[0]))
#     series = Series(values, xvalues, title_from_data=True)
#     c2.series.append(series)
#
#     s2 = c2.series[0]
#     s2.marker.symbol = "diamond"
#     s2.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
#     s2.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
#     s2.graphicalProperties.line.noFill = True  # hide lines
#
#     c2.style = 13
#     # c2.x_axis.title = 'Days'
#     c2.y_axis.title = 'S, мм/с'
#     c2.y_axis.axId = 200
#     c2.y_axis.majorGridlines = None
#     c2.x_axis.majorGridlines = None
#
#     # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
#     c1.y_axis.crosses = "max"
#     c1 += c2
#
#     sheet.add_chart(c1, "O10")
#
#     book.save("output_data/fd.xlsx")


# example()
